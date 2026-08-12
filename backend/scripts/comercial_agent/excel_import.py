"""
Lectura y normalización del Excel "POLO 52 - EMPRESAS INSTALADAS".

La hoja "2026" es el listado vigente de empresas (nombre, actividad, contacto,
ubicación, cantidad de empleados), pero no trae CUIL. La hoja "Empresas 2025"
sí trae CUIL para una parte de las empresas, así que se cruza por nombre
normalizado para completarlo cuando existe. Para las que no tienen CUIL en
ninguna de las dos hojas se genera un CUIL placeholder claramente falso
(prefijo "99", que ningún CUIL real argentino usa) para no romper la clave
primaria de `empresa`; se marca `cuil_placeholder=True` para que quede visible
en la revisión y no se confunda con un dato real.
"""
import re
import unicodedata
from dataclasses import dataclass, field
from typing import Dict, List, Optional

import openpyxl

SHEET_ACTUAL = "2026"
SHEET_CUILS = "Empresas 2025"

_PLACEHOLDER_CUIL_PREFIX = 99_000_000_000  # ningún CUIL real argentino empieza con 99


def _normalize_name(value: Optional[str]) -> str:
    if not value:
        return ""
    normalized = "".join(
        c for c in unicodedata.normalize("NFD", str(value)) if unicodedata.category(c) != "Mn"
    )
    normalized = re.sub(r"[^a-z0-9 ]", "", normalized.lower())
    return re.sub(r"\s+", " ", normalized).strip()


@dataclass
class EmpresaExcelRow:
    nombre: str
    rubro: Optional[str] = None
    contacto_nombre: Optional[str] = None
    contacto_telefono: Optional[str] = None
    cant_empleados: Optional[int] = None
    pagina_web: Optional[str] = None
    cuil: Optional[int] = None
    cuil_placeholder: bool = False
    tipologia: Optional[str] = None
    ubicacion_raw: Optional[str] = None
    propietario: Optional[str] = None
    fecha_ingreso: Optional[str] = None  # ISO string si el excel la trae
    notas_excel: List[str] = field(default_factory=list)


def _load_cuiles_por_nombre(path: str) -> Dict[str, int]:
    wb = openpyxl.load_workbook(path, data_only=True)
    if SHEET_CUILS not in wb.sheetnames:
        return {}
    ws = wb[SHEET_CUILS]
    headers = [c.value for c in ws[1]]
    idx_empresa = headers.index("Empresa")
    idx_cuil = headers.index("CUIL")

    result: Dict[str, int] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        nombre = row[idx_empresa]
        cuil_raw = row[idx_cuil]
        if not nombre or not cuil_raw:
            continue
        digits = re.sub(r"\D", "", str(cuil_raw))
        if len(digits) != 11:
            continue
        result[_normalize_name(nombre)] = int(digits)
    return result


def _parse_telefono(value) -> Optional[str]:
    if value in (None, ""):
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _parse_cant_empleados(value) -> Optional[int]:
    if value in (None, ""):
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def _parse_fecha_ingreso(value) -> Optional[str]:
    if value in (None, ""):
        return None
    import datetime as _dt

    if isinstance(value, (_dt.date, _dt.datetime)):
        return value.date().isoformat() if isinstance(value, _dt.datetime) else value.isoformat()
    return None  # texto libre no confiable como fecha, se ignora


def load_empresas(path: str) -> List[EmpresaExcelRow]:
    """Devuelve la lista normalizada de empresas de la hoja vigente ("2026")."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[SHEET_ACTUAL]
    headers = [c.value for c in ws[1]]

    def idx(col_name: str) -> int:
        return headers.index(col_name)

    idx_empresa = idx("Empresa")
    idx_actividad = idx("Actividad")
    idx_contacto_nombre = idx("Nombre del contacto")
    idx_contacto_tel = idx("Contacto Inquilino")
    idx_empleados = idx("Cantidad de Empleados")
    idx_web = idx("Página web")
    idx_tipologia = idx("Tipología (Nave-Oficina-Local Comercial-Coworking-Parking)")
    idx_ubicacion = idx("Ubicación")
    idx_propietario = idx("Propietario")
    idx_fecha = idx("Fecha de instalación")

    cuiles_por_nombre = _load_cuiles_por_nombre(path)

    empresas: List[EmpresaExcelRow] = []
    seen_placeholder = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        nombre = row[idx_empresa]
        if not nombre or not str(nombre).strip():
            continue
        nombre = str(nombre).strip()

        cuil = cuiles_por_nombre.get(_normalize_name(nombre))
        cuil_placeholder = False
        notas = []
        if not cuil:
            cuil = _PLACEHOLDER_CUIL_PREFIX + seen_placeholder
            seen_placeholder += 1
            cuil_placeholder = True
            notas.append(
                "CUIL placeholder generado automaticamente: no estaba cargado en el excel. "
                "Reemplazar por el CUIL real antes de considerar el registro definitivo."
            )

        empresas.append(
            EmpresaExcelRow(
                nombre=nombre,
                rubro=(str(row[idx_actividad]).strip() if row[idx_actividad] else None),
                contacto_nombre=(str(row[idx_contacto_nombre]).strip() if row[idx_contacto_nombre] else None),
                contacto_telefono=_parse_telefono(row[idx_contacto_tel]),
                cant_empleados=_parse_cant_empleados(row[idx_empleados]),
                pagina_web=(str(row[idx_web]).strip() if row[idx_web] else None),
                cuil=cuil,
                cuil_placeholder=cuil_placeholder,
                tipologia=(str(row[idx_tipologia]).strip() if row[idx_tipologia] else None),
                ubicacion_raw=(str(row[idx_ubicacion]).strip() if row[idx_ubicacion] else None),
                propietario=(str(row[idx_propietario]).strip() if row[idx_propietario] else None),
                fecha_ingreso=_parse_fecha_ingreso(row[idx_fecha]),
                notas_excel=notas,
            )
        )

    return empresas


if __name__ == "__main__":
    import sys

    path = sys.argv[1] if len(sys.argv) > 1 else (
        "/Users/maxiquelas/Downloads/POLO 52 - EMPRESAS INSTALADAS 2026.xlsx"
    )
    rows = load_empresas(path)
    print(f"{len(rows)} empresas cargadas del excel")
    con_web = sum(1 for r in rows if r.pagina_web)
    con_cuil_real = sum(1 for r in rows if not r.cuil_placeholder)
    print(f"  con pagina web en el excel: {con_web}")
    print(f"  con CUIL real (cruzado con hoja {SHEET_CUILS!r}): {con_cuil_real}")
    for r in rows[:5]:
        print(" ", r)
